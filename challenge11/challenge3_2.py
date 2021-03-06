from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

def combine():
    wb = load_workbook('courses.xlsx')
    sheet_students = wb['students']
    sheet_time = wb['time']
    ws2 = wb.copy_worksheet(sheet_students)
    ws2.title = 'combine'
    ws2['D1'] = sheet_time['C1'].value

    data = {}

    for row in sheet_students.rows:
        if row[1].value == sheet_students['B1'].value: 
            pass
        else:
            data[row[1].value] = [row[0].value, row[2].value]

    for row in sheet_time.rows:
        if row[1].value == sheet_students['B1'].value:
            pass
        else:
            data[row[1].value].append(row[2].value)

    for i in range(2, sheet_students.max_row+1):
        ws2.cell(row=i, column=4).value = data[ws2.cell(row=i, column=2).value][2]

    wb.save("courses.xlsx")


def split():
    wb = load_workbook('courses.xlsx')
    sheet_c = wb['combine']
    data = {}
    title = [sheet_c['A1'].value, sheet_c['B1'].value, sheet_c['C1'].value, sheet_c['D1'].value]
    for i in range(2, sheet_c.max_row+1):
        date = sheet_c.cell(row=i, column=1).value
        year = date.year
        data[year] = data.get(year, [1,])
        data_value = []
        for c in list(sheet_c.rows)[i-1]:
            data_value.append(c.value)
        data[year].append(data_value)
#    print(data)
 
    for key, value in data.items():
        ws = Workbook()
        ws1 = ws.active
        ws1.title = str(key)
        ws1.append(title)
        for i in value[1:]:
            ws1.append(i)
        for i in range(2, ws1.max_row+1):
            ws1.cell(row=i, column=1).number_format = 'yyyy/m/d h:mm'
        ws.save(str(key)+'.xlsx')

if __name__ == '__main__':
    combine()
    split()

